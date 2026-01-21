# import packages and all necessary modules

import openpyxl
from openpyxl import load_workbook
from openpyxl.descriptors import (String,Sequence,Integer)
from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.styles import numbers,Alignment,PatternFill,Font,colors
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
import numpy as np, pandas as pd
from pandas import DataFrame
import copy
from datetime import datetime
from datetime import date
from dateutil.relativedelta import relativedelta
import warnings

today=input("what day is today? (yyyy,mm,dd)")
today=datetime.strptime(today,"%Y,%m,%d")
week=today.strftime("%V")
yr=today.strftime("%Y")
wk=yr+week
mon=today.strftime("%b")
next_month=today+relativedelta(months=+1)
next_mon=next_month.strftime("%b")


warnings.filterwarnings('ignore',category=UserWarning)

def clean(ws,maxcol):
    for col in ws.iter_cols(max_col=maxcol,max_row=len(ws['B'])):
        for cell in col:
            cell.value=''
            cell.font = NamedStyle().font
            cell.fill = NamedStyle().fill
            cell.border = NamedStyle().border
            cell.alignment = NamedStyle().alignment
            cell.number_format = NamedStyle().number_format

# data_col: last data column，数据的最后一行，后边可能是pivot
# arrive_push: PSI arrive, if push the last week to next month, default is YES
# piv: 0 means no pivot table in the sheet, 1 means yes and refresh, default Yes

def gscm(target, gfile, tsheet, data_col, arrive_push=1, piv=1):
    with pd.ExcelWriter(target, mode='a', engine='openpyxl', if_sheet_exists="overlay") as writer:
        workbook = writer.book
        w1 = writer.sheets[tsheet]
        w2_wb = load_workbook(gfile)
        w2 = w2_wb.active

        clean(w1, data_col)

        found_maxrow = 0
        for rows in w1.iter_rows(min_col=5,max_col=5):
            for cell in rows:
                if cell.value is None:
                    maxrow=cell.row
                    found_maxrow = 1
            if found_maxrow == 1:
                break

        for row in w2.iter_rows(min_row=2, min_col=2):
            for cell in row:
                nc = w1.cell(row=cell.row - 1, column=cell.column - 1, value=cell.value)
                if cell.has_style:
                    nc.font = copy.copy(cell.font)
                    nc.fill = copy.copy(cell.fill)
                    nc.border = copy.copy(cell.border)
                    nc.alignment = copy.copy(cell.alignment)
                    nc.number_format = copy.copy(cell.number_format)
        mon_idx=0
        nondate_idx=0
        next_mon_idx=0
        wk_idx=0
        if tsheet == "PSI Arrivals":
            for toprow in w1.iter_rows(min_row=1, max_row=1, max_col=data_col):
                    for header in toprow:
                        if header.value == mon:
                            mon_idx = header.col_idx
                        elif header.value == "Category":
                            nondate_idx = header.col_idx
                        elif header.value == next_mon:
                            next_mon_idx = header.col_idx
                        elif header.value == wk:
                            wk_idx = header.col_idx
                        elif wk_idx != 0 and next_mon_idx != 0 and nondate_idx != 0 and mon_idx != 0:
                            break
            wk_qty = mon_idx - wk_idx
            nxwk_qty = next_mon_idx - mon_idx

            for i in range(3,maxrow):
                j = 0
                k = 1
                cur_total = 0
                nx_total = 0
                if arrive_push != 1:
                    while j < wk_qty:
                        wk_select = wk_idx + j
                        cur_total = cur_total + w1.cell(i, wk_select).value
                        j += 1
                else:
                    while j < wk_qty - 1:
                        wk_select = wk_idx + j
                        cur_total = cur_total + w1.cell(i, wk_select).value
                        j += 1
                    nx_total = nx_total + w1.cell(i, mon_idx - 1).value
                    while k < nxwk_qty:
                        wk_select = mon_idx + k
                        nx_total = nx_total + w1.cell(i, wk_select).value
                        k += 1
                    w1.cell(i, next_mon_idx).value = nx_total
                w1.cell(i, mon_idx).value = cur_total


        elif tsheet=="AP1":
            for toprow in w1.iter_rows(min_row=1, max_row=1, max_col=data_col):
                    for header in toprow:
                        if header.value == mon:
                            mon_idx = header.col_idx
                        elif header.value == "Category":
                            nondate_idx = header.col_idx
                        elif header.value == wk:
                            wk_idx = header.col_idx
                        elif wk_idx != 0 and nondate_idx != 0 and mon_idx != 0:
                            break
            wk_qty = mon_idx - wk_idx
            for i in range(7,maxrow):
                # 5 个 一循环
                j = 0
                cur_total = 0
                if (i-6) % 5 !=2 :
                    while j < wk_qty:
                        wk_select = wk_idx + j
                        cur_total = cur_total + w1.cell(i, wk_select).value
                        j += 1
                    w1.cell(i, mon_idx).value = cur_total
                else:
                    cur_total = cur_total + w1.cell(i+1, wk_select).value
                    j=1
                    while j < wk_qty:
                        wk_select = wk_idx + j
                        cur_total = cur_total + w1.cell(i, wk_select).value
                        j += 1

        if piv == 1:
            pivot = w1._pivots[0]
            pivot.cache.refreshOnLoad = True


def sap(target, sfile, tsheet,data_col):
    with pd.ExcelWriter(target, mode='a', engine='openpyxl', if_sheet_exists="overlay") as writer:
        workbook = writer.book
        w1 = writer.sheets[tsheet]
        w2_wb = load_workbook(sfile)
        w2 = w2_wb.active

        clean(w1, data_col)

        for row in w2.iter_rows():
            for cell in row:
                nc = w1.cell(row=cell.row, column=cell.column, value=cell.value)


gscm("E:\\三星\\ATS\\wk48\\python.xlsx",
     "E:\\三星\\ATS\\wk48\\arrive.xlsx","PSI Arrivals",42)
